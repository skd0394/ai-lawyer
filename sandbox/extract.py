"""Document extraction — runs INSIDE the sandbox.

Every format converges on MARKDOWN WITH HEADING LEVELS PRESERVED. That is
the load-bearing decision: headings are what make mode="section" possible.
Flatten to plain text and "read the termination clause" degenerates into
"read the whole document", which is how context budgets die.

Modes:
    outline   structure only, ~200 tokens. The default.
    section   one named section
    full      everything, capped

All heavy imports are lazy so the API container can import this module
without pymupdf installed.
"""

from __future__ import annotations

import base64
import csv
import io
import os
import re

# A scanned page yields near-zero extractable text. Below this many
# characters per page, treat it as an image and send it for vision OCR.
OCR_CHARS_PER_PAGE = 100
MAX_OCR_PAGES = 8           # cap: vision OCR is slow and costs tokens
OCR_DPI = 200               # below ~150 accuracy drops sharply


# ── helpers ───────────────────────────────────────────────────────────────
def _headings(md: str) -> list[str]:
    return [ln.strip() for ln in md.splitlines() if ln.strip().startswith("#")]


def _section(md: str, wanted: str) -> str:
    """Return the named heading and everything until the next heading of the
    same or higher level."""
    lines = md.splitlines()
    want = wanted.strip().lstrip("#").strip().lower()
    start = level = None
    for i, ln in enumerate(lines):
        st = ln.strip()
        if st.startswith("#") and want in st.lstrip("#").strip().lower():
            start = i
            level = len(st) - len(st.lstrip("#"))
            break
    if start is None:
        return ""
    out = [lines[start]]
    for ln in lines[start + 1:]:
        st = ln.strip()
        if st.startswith("#"):
            if (len(st) - len(st.lstrip("#"))) <= level:
                break
        out.append(ln)
    return "\n".join(out)


def _cap(text: str, limit: int) -> tuple[str, bool]:
    if len(text) <= limit:
        return text, False
    return text[:limit], True


# ── PDF ───────────────────────────────────────────────────────────────────
def _pdf(path: str, mode: str, want_images: bool) -> dict:
    import fitz                                   # pymupdf

    doc = fitz.open(path)
    pages = [p.get_text() for p in doc]
    total_chars = sum(len(t.strip()) for t in pages)
    scanned = total_chars < OCR_CHARS_PER_PAGE * max(len(pages), 1)

    if scanned:
        out = {"kind": "pdf", "pages": len(doc), "needs_ocr": True,
               "outline": [f"(scanned document, {len(doc)} pages)"],
               "text": "", "total_chars": total_chars}
        if want_images:
            imgs = []
            for p in doc[:MAX_OCR_PAGES]:
                pix = p.get_pixmap(dpi=OCR_DPI)
                imgs.append(base64.b64encode(pix.tobytes("png")).decode())
            out["page_images_b64"] = imgs
            out["ocr_pages_included"] = len(imgs)
            out["ocr_pages_omitted"] = max(0, len(doc) - MAX_OCR_PAGES)
        doc.close()
        return out

    # Text PDF. Use the embedded table of contents when present; otherwise
    # fall back to page markers so section reads still have anchors.
    toc = []
    try:
        toc = doc.get_toc() or []
    except Exception:
        pass

    parts = []
    if toc:
        for lvl, title, page in toc:
            parts.append(f"{'#' * min(lvl + 1, 6)} {title}")
        body = "\n\n".join(f"## Page {i+1}\n\n{t.strip()}"
                           for i, t in enumerate(pages) if t.strip())
        md = "\n".join(parts) + "\n\n" + body
        outline = [f"{'  ' * (l-1)}{t} (p{p})" for l, t, p in toc]
    else:
        md = "\n\n".join(f"## Page {i+1}\n\n{t.strip()}"
                         for i, t in enumerate(pages) if t.strip())
        outline = [f"Page {i+1}: {(t.strip().splitlines() or [''])[0][:70]}"
                   for i, t in enumerate(pages) if t.strip()]

    doc.close()
    return {"kind": "pdf", "pages": len(pages), "needs_ocr": False,
            "outline": outline, "text": md, "total_chars": len(md)}


# ── DOCX ──────────────────────────────────────────────────────────────────
def _docx(path: str) -> dict:
    from docx import Document

    doc = Document(path)
    lines: list[str] = []

    for block in doc.element.body.iterchildren():
        tag = block.tag.split("}")[-1]

        if tag == "p":
            from docx.text.paragraph import Paragraph
            p = Paragraph(block, doc)
            text = p.text.strip()
            if not text:
                continue
            style = (p.style.name or "").lower()
            if style.startswith("heading"):
                m = re.search(r"(\d+)", style)
                lvl = min(int(m.group(1)) if m else 1, 6)
                lines.append(f"{'#' * lvl} {text}")
            elif style.startswith("title"):
                lines.append(f"# {text}")
            elif "list" in style:
                lines.append(f"- {text}")
            else:
                lines.append(text)

        elif tag == "tbl":
            from docx.table import Table
            t = Table(block, doc)
            rows = [[c.text.strip().replace("|", "\\|") for c in r.cells]
                    for r in t.rows]
            if not rows:
                continue
            lines.append("| " + " | ".join(rows[0]) + " |")
            lines.append("|" + "---|" * len(rows[0]))
            for r in rows[1:]:
                lines.append("| " + " | ".join(r) + " |")
            lines.append("")

    md = "\n\n".join(lines)
    heads = _headings(md)
    return {"kind": "docx", "needs_ocr": False,
            "outline": heads or [f"(no headings, {len(lines)} blocks)"],
            "text": md, "total_chars": len(md),
            "paragraphs": len(lines)}


# ── spreadsheets ──────────────────────────────────────────────────────────
def _xlsx(path: str, max_rows: int = 30) -> dict:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    parts, outline, sheets = [], [], []

    for ws in wb.worksheets:
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                break
            rows.append(["" if c is None else str(c) for c in row])

        total = ws.max_row or 0
        sheets.append({"name": ws.title, "rows": total, "cols": ws.max_column})
        outline.append(f"Sheet '{ws.title}' ({total} rows x {ws.max_column} cols)")
        parts.append(f"## Sheet: {ws.title}\n"
                     f"({total} rows x {ws.max_column} columns)\n")

        if rows:
            width = max(len(r) for r in rows)
            rows = [r + [""] * (width - len(r)) for r in rows]
            parts.append("| " + " | ".join(rows[0]) + " |")
            parts.append("|" + "---|" * width)
            for r in rows[1:]:
                parts.append("| " + " | ".join(r) + " |")
            # Never dump an entire sheet — say what was withheld instead.
            if total > max_rows:
                parts.append(f"\n[{total - max_rows} more rows not shown]")
        parts.append("")

    wb.close()
    md = "\n".join(parts)
    return {"kind": "xlsx", "needs_ocr": False, "outline": outline,
            "text": md, "total_chars": len(md), "sheets": sheets}


def _csv(path: str, max_rows: int = 30) -> dict:
    with open(path, newline="", encoding="utf-8", errors="replace") as f:
        rows = list(csv.reader(f))
    if not rows:
        return {"kind": "csv", "needs_ocr": False, "outline": ["(empty)"],
                "text": "", "total_chars": 0}
    width = max(len(r) for r in rows[:max_rows + 1])
    shown = [r + [""] * (width - len(r)) for r in rows[:max_rows + 1]]
    parts = ["| " + " | ".join(shown[0]) + " |", "|" + "---|" * width]
    for r in shown[1:]:
        parts.append("| " + " | ".join(r) + " |")
    if len(rows) > max_rows + 1:
        parts.append(f"\n[{len(rows) - max_rows - 1} more rows not shown]")
    md = "\n".join(parts)
    return {"kind": "csv", "needs_ocr": False,
            "outline": [f"CSV: {len(rows)} rows x {width} columns",
                        "Columns: " + ", ".join(shown[0])],
            "text": md, "total_chars": len(md), "rows": len(rows)}


# ── plain text / images ───────────────────────────────────────────────────
def _text(path: str) -> dict:
    with open(path, encoding="utf-8", errors="replace") as f:
        md = f.read()
    heads = _headings(md)
    return {"kind": "text", "needs_ocr": False,
            "outline": heads or [f"({len(md)} chars, "
                                 f"{len(md.splitlines())} lines)"],
            "text": md, "total_chars": len(md)}


def _image(path: str) -> dict:
    """Images go to the model natively — there is nothing to extract."""
    ext = os.path.splitext(path)[1].lower().lstrip(".")
    mime = {"jpg": "jpeg", "jpeg": "jpeg", "png": "png",
            "webp": "webp", "gif": "gif"}.get(ext, "png")
    with open(path, "rb") as f:
        raw = f.read()
    return {"kind": "image", "needs_ocr": True, "outline": ["(image)"],
            "text": "", "total_chars": 0,
            "image_b64": base64.b64encode(raw).decode(),
            "image_mime": f"image/{mime}", "bytes": len(raw)}


# ── entry point ───────────────────────────────────────────────────────────
def extract(path: str, mode: str = "outline", section: str | None = None,
            max_chars: int = 8000, want_images: bool = False) -> dict:
    ext = os.path.splitext(path)[1].lower()

    if ext == ".pdf":
        r = _pdf(path, mode, want_images)
    elif ext in (".docx", ".doc"):
        r = _docx(path)
    elif ext in (".xlsx", ".xls"):
        r = _xlsx(path)
    elif ext == ".csv":
        r = _csv(path)
    elif ext in (".png", ".jpg", ".jpeg", ".webp", ".gif"):
        r = _image(path)
    else:
        r = _text(path)

    r["mode"] = mode
    r["file"] = os.path.basename(path)

    if r.get("needs_ocr") or not r.get("text"):
        return r

    if mode == "outline":
        r["text"] = ""
        return r

    if mode == "section":
        if not section:
            r["text"] = ""
            r["error"] = "mode='section' requires a section name"
            return r
        body = _section(r["text"], section)
        if not body:
            r["text"] = ""
            r["error"] = (f"no section matching '{section}'. "
                          f"Available: {r.get('outline', [])[:20]}")
            return r
        r["text"] = body

    r["text"], r["truncated"] = _cap(r["text"], max_chars)
    return r